from __future__ import annotations

from collections.abc import Callable, Iterator
from contextlib import contextmanager
from copy import copy
from pathlib import Path
import re
from typing import Literal, Protocol, cast, runtime_checkable

from pydantic import BaseModel, Field, field_validator, model_validator
import xlwings as xw

from exstruct.cli.availability import get_com_availability

from .extract_runner import OnConflictPolicy
from .io import PathPolicy

PatchOpType = Literal[
    "set_value",
    "set_formula",
    "add_sheet",
    "set_range_values",
    "fill_formula",
    "set_value_if",
    "set_formula_if",
    "draw_grid_border",
    "set_bold",
    "set_fill_color",
    "set_dimensions",
    "merge_cells",
    "unmerge_cells",
    "set_alignment",
    "restore_design_snapshot",
]
PatchStatus = Literal["applied", "skipped"]
PatchValueKind = Literal["value", "formula", "sheet", "style", "dimension"]
FormulaIssueLevel = Literal["warning", "error"]
FormulaIssueCode = Literal[
    "invalid_token",
    "ref_error",
    "name_error",
    "div0_error",
    "value_error",
    "na_error",
    "circular_ref_suspected",
]

_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
_A1_PATTERN = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")
_A1_RANGE_PATTERN = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*:[A-Za-z]{1,3}[1-9][0-9]*$")
_HEX_COLOR_PATTERN = re.compile(r"^#(?:[0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})$")
_COLUMN_LABEL_PATTERN = re.compile(r"^[A-Za-z]{1,3}$")
_MAX_STYLE_TARGET_CELLS = 10_000

HorizontalAlignType = Literal[
    "general",
    "left",
    "center",
    "right",
    "fill",
    "justify",
    "centerContinuous",
    "distributed",
]
VerticalAlignType = Literal["top", "center", "bottom", "justify", "distributed"]


class BorderSideSnapshot(BaseModel):
    """Serializable border side state for inverse restoration."""

    style: str | None = None
    color: str | None = None


class BorderSnapshot(BaseModel):
    """Serializable border state for one cell."""

    cell: str
    top: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)
    right: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)
    bottom: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)
    left: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)


class FontSnapshot(BaseModel):
    """Serializable font state for one cell."""

    cell: str
    bold: bool | None = None


class FillSnapshot(BaseModel):
    """Serializable fill state for one cell."""

    cell: str
    fill_type: str | None = None
    start_color: str | None = None
    end_color: str | None = None


class AlignmentSnapshot(BaseModel):
    """Serializable alignment state for one cell."""

    cell: str
    horizontal: str | None = None
    vertical: str | None = None
    wrap_text: bool | None = None


class MergeStateSnapshot(BaseModel):
    """Serializable merged-range state for deterministic restoration."""

    scope: str
    ranges: list[str] = Field(default_factory=list)


class RowDimensionSnapshot(BaseModel):
    """Serializable row height state."""

    row: int
    height: float | None = None


class ColumnDimensionSnapshot(BaseModel):
    """Serializable column width state."""

    column: str
    width: float | None = None


class DesignSnapshot(BaseModel):
    """Serializable style/dimension snapshot for inverse restore."""

    borders: list[BorderSnapshot] = Field(default_factory=list)
    fonts: list[FontSnapshot] = Field(default_factory=list)
    fills: list[FillSnapshot] = Field(default_factory=list)
    alignments: list[AlignmentSnapshot] = Field(default_factory=list)
    merge_state: MergeStateSnapshot | None = None
    row_dimensions: list[RowDimensionSnapshot] = Field(default_factory=list)
    column_dimensions: list[ColumnDimensionSnapshot] = Field(default_factory=list)


@runtime_checkable
class OpenpyxlCellProtocol(Protocol):
    """Protocol for openpyxl cell access used by patch runner."""

    value: str | int | float | None
    data_type: str | None
    font: OpenpyxlFontProtocol
    fill: OpenpyxlFillProtocol
    border: OpenpyxlBorderProtocol
    alignment: OpenpyxlAlignmentProtocol


@runtime_checkable
class OpenpyxlColorProtocol(Protocol):
    """Protocol for openpyxl color access."""

    rgb: object | None


@runtime_checkable
class OpenpyxlSideProtocol(Protocol):
    """Protocol for openpyxl border side access."""

    style: str | None
    color: OpenpyxlColorProtocol | None


@runtime_checkable
class OpenpyxlBorderProtocol(Protocol):
    """Protocol for openpyxl border access."""

    top: OpenpyxlSideProtocol
    right: OpenpyxlSideProtocol
    bottom: OpenpyxlSideProtocol
    left: OpenpyxlSideProtocol


@runtime_checkable
class OpenpyxlFontProtocol(Protocol):
    """Protocol for openpyxl font access."""

    bold: bool | None


@runtime_checkable
class OpenpyxlFillProtocol(Protocol):
    """Protocol for openpyxl fill access."""

    fill_type: str | None
    start_color: OpenpyxlColorProtocol | None
    end_color: OpenpyxlColorProtocol | None


@runtime_checkable
class OpenpyxlAlignmentProtocol(Protocol):
    """Protocol for openpyxl alignment access."""

    horizontal: str | None
    vertical: str | None
    wrap_text: bool | None


@runtime_checkable
class OpenpyxlRowDimensionProtocol(Protocol):
    """Protocol for openpyxl row dimension access."""

    height: float | None


@runtime_checkable
class OpenpyxlColumnDimensionProtocol(Protocol):
    """Protocol for openpyxl column dimension access."""

    width: float | None


@runtime_checkable
class OpenpyxlRowDimensionsProtocol(Protocol):
    """Protocol for openpyxl row dimensions collection."""

    def __getitem__(self, key: int) -> OpenpyxlRowDimensionProtocol: ...


@runtime_checkable
class OpenpyxlColumnDimensionsProtocol(Protocol):
    """Protocol for openpyxl column dimensions collection."""

    def __getitem__(self, key: str) -> OpenpyxlColumnDimensionProtocol: ...


@runtime_checkable
class OpenpyxlWorksheetProtocol(Protocol):
    """Protocol for openpyxl worksheet access used by patch runner."""

    row_dimensions: OpenpyxlRowDimensionsProtocol
    column_dimensions: OpenpyxlColumnDimensionsProtocol

    def __getitem__(self, key: str) -> OpenpyxlCellProtocol: ...

    def merge_cells(self, range_string: str) -> None: ...

    def unmerge_cells(self, range_string: str) -> None: ...


@runtime_checkable
class OpenpyxlWorkbookProtocol(Protocol):
    """Protocol for openpyxl workbook access used by patch runner."""

    sheetnames: list[str]

    def __getitem__(self, key: str) -> OpenpyxlWorksheetProtocol: ...

    def create_sheet(self, title: str) -> OpenpyxlWorksheetProtocol: ...

    def save(self, filename: str | Path) -> None: ...

    def close(self) -> None: ...


@runtime_checkable
class XlwingsRangeProtocol(Protocol):
    """Protocol for xlwings range access used by patch runner."""

    value: str | int | float | None
    formula: str | None


@runtime_checkable
class XlwingsSheetProtocol(Protocol):
    """Protocol for xlwings sheet access used by patch runner."""

    name: str

    def range(self, cell: str) -> XlwingsRangeProtocol: ...


@runtime_checkable
class XlwingsSheetsProtocol(Protocol):
    """Protocol for xlwings sheets collection."""

    def __iter__(self) -> Iterator[XlwingsSheetProtocol]: ...

    def __len__(self) -> int: ...

    def __getitem__(self, index: int) -> XlwingsSheetProtocol: ...

    def add(
        self, name: str, after: XlwingsSheetProtocol | None = None
    ) -> XlwingsSheetProtocol: ...


@runtime_checkable
class XlwingsWorkbookProtocol(Protocol):
    """Protocol for xlwings workbook access used by patch runner."""

    sheets: XlwingsSheetsProtocol

    def save(self, filename: str) -> None: ...

    def close(self) -> None: ...


class PatchOp(BaseModel):
    """Single patch operation for an Excel workbook.

    Operation types and their required fields:

    - ``set_value``: Set a cell value. Requires ``sheet``, ``cell``, ``value``.
    - ``set_formula``: Set a cell formula. Requires ``sheet``, ``cell``, ``formula`` (must start with ``=``).
    - ``add_sheet``: Add a new worksheet. Requires ``sheet`` (new sheet name). No ``cell``/``value``/``formula``.
    - ``set_range_values``: Set values for a rectangular range. Requires ``sheet``, ``range`` (e.g. ``A1:C3``), ``values`` (2D list matching range shape).
    - ``fill_formula``: Fill a formula across a single row or column. Requires ``sheet``, ``range``, ``base_cell``, ``formula``.
    - ``set_value_if``: Conditionally set value. Requires ``sheet``, ``cell``, ``value``. ``expected`` is optional; ``null`` matches an empty cell. Skips if current value != expected.
    - ``set_formula_if``: Conditionally set formula. Requires ``sheet``, ``cell``, ``formula``. ``expected`` is optional; ``null`` matches an empty cell. Skips if current value != expected.
    - ``draw_grid_border``: Draw thin black borders on a target rectangle.
    - ``set_bold``: Set bold style for one cell or one range.
    - ``set_fill_color``: Set solid fill color for one cell or one range.
    - ``set_dimensions``: Set row height and/or column width.
    - ``merge_cells``: Merge a rectangular range.
    - ``unmerge_cells``: Unmerge all merged ranges intersecting target range.
    - ``set_alignment``: Set horizontal/vertical alignment and/or wrap_text.
    - ``restore_design_snapshot``: Restore style/dimension snapshot (internal inverse op).
    """

    op: PatchOpType = Field(
        description=(
            "Operation type: 'set_value', 'set_formula', 'add_sheet', "
            "'set_range_values', 'fill_formula', 'set_value_if', 'set_formula_if', "
            "'draw_grid_border', 'set_bold', 'set_fill_color', 'set_dimensions', "
            "'merge_cells', 'unmerge_cells', 'set_alignment', "
            "or 'restore_design_snapshot'."
        )
    )
    sheet: str = Field(
        description="Target sheet name. For add_sheet, this is the new sheet name."
    )
    cell: str | None = Field(
        default=None,
        description="Cell reference in A1 notation (e.g. 'B2'). Required for set_value, set_formula, set_value_if, set_formula_if.",
    )
    range: str | None = Field(
        default=None,
        description="Range reference in A1 notation (e.g. 'A1:C3'). Required for set_range_values and fill_formula.",
    )
    base_cell: str | None = Field(
        default=None,
        description="Base cell for formula translation in fill_formula (e.g. 'C2').",
    )
    expected: str | int | float | None = Field(
        default=None,
        description="Expected current value for conditional ops (set_value_if, set_formula_if). Operation is skipped if mismatch.",
    )
    value: str | int | float | None = Field(
        default=None,
        description="Value to set. Use null to clear a cell. For set_value and set_value_if.",
    )
    values: list[list[str | int | float | None]] | None = Field(
        default=None,
        description="2D list of values for set_range_values. Shape must match the range dimensions.",
    )
    formula: str | None = Field(
        default=None,
        description="Formula string starting with '=' (e.g. '=SUM(A1:A10)'). For set_formula, set_formula_if, fill_formula.",
    )
    row_count: int | None = Field(
        default=None,
        description="Row count for draw_grid_border.",
    )
    col_count: int | None = Field(
        default=None,
        description="Column count for draw_grid_border.",
    )
    bold: bool | None = Field(
        default=None,
        description="Bold flag for set_bold. Defaults to true.",
    )
    fill_color: str | None = Field(
        default=None,
        description="Fill color for set_fill_color in #RRGGBB or #AARRGGBB format.",
    )
    rows: list[int] | None = Field(
        default=None,
        description="Row indexes for set_dimensions.",
    )
    columns: list[str | int] | None = Field(
        default=None,
        description="Column identifiers for set_dimensions. Accepts letters (A/AA) or positive indexes.",
    )
    row_height: float | None = Field(
        default=None,
        description="Target row height for set_dimensions.",
    )
    column_width: float | None = Field(
        default=None,
        description="Target column width for set_dimensions.",
    )
    horizontal_align: HorizontalAlignType | None = Field(
        default=None,
        description="Horizontal alignment for set_alignment.",
    )
    vertical_align: VerticalAlignType | None = Field(
        default=None,
        description="Vertical alignment for set_alignment.",
    )
    wrap_text: bool | None = Field(
        default=None,
        description="Wrap text flag for set_alignment.",
    )
    design_snapshot: DesignSnapshot | None = Field(
        default=None,
        description="Design snapshot payload for restore_design_snapshot.",
    )

    @field_validator("sheet")
    @classmethod
    def _validate_sheet(cls, value: str) -> str:
        if not value.strip():
            raise ValueError("sheet must not be empty.")
        return value

    @field_validator("cell")
    @classmethod
    def _validate_cell(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not _A1_PATTERN.match(candidate):
            raise ValueError(f"Invalid cell reference: {value}")
        return candidate.upper()

    @field_validator("base_cell")
    @classmethod
    def _validate_base_cell(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not _A1_PATTERN.match(candidate):
            raise ValueError(f"Invalid base_cell reference: {value}")
        return candidate.upper()

    @field_validator("range")
    @classmethod
    def _validate_range(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not _A1_RANGE_PATTERN.match(candidate):
            raise ValueError(f"Invalid range reference: {value}")
        start, end = candidate.split(":", maxsplit=1)
        return f"{start.upper()}:{end.upper()}"

    @field_validator("fill_color")
    @classmethod
    def _validate_fill_color(cls, value: str | None) -> str | None:
        if value is None:
            return None
        if not _HEX_COLOR_PATTERN.match(value):
            raise ValueError("Invalid fill_color format. Use '#RRGGBB' or '#AARRGGBB'.")
        return value.upper()

    @field_validator("rows")
    @classmethod
    def _validate_rows(cls, value: list[int] | None) -> list[int] | None:
        if value is None:
            return None
        if not value:
            raise ValueError("rows must not be empty.")
        normalized: list[int] = []
        for row in value:
            if row < 1:
                raise ValueError("rows must contain positive integers.")
            normalized.append(row)
        return normalized

    @field_validator("columns")
    @classmethod
    def _validate_columns(cls, value: list[str | int] | None) -> list[str | int] | None:
        if value is None:
            return None
        if not value:
            raise ValueError("columns must not be empty.")
        normalized: list[str | int] = []
        for column in value:
            normalized.append(_normalize_column_identifier(column))
        return normalized

    @model_validator(mode="after")
    def _validate_op(self) -> PatchOp:
        validator = _validator_for_op(self.op)
        if validator is None:
            return self
        if self.op in _CELL_REQUIRED_OPS:
            _validate_cell_required(self)
        validator(self)
        return self


_CELL_REQUIRED_OPS: set[PatchOpType] = {
    "set_value",
    "set_formula",
    "set_value_if",
    "set_formula_if",
}


def _validator_for_op(op_type: PatchOpType) -> Callable[[PatchOp], None] | None:
    """Return per-op validator function."""
    validators: dict[PatchOpType, Callable[[PatchOp], None]] = {
        "add_sheet": _validate_add_sheet,
        "set_value": _validate_set_value,
        "set_formula": _validate_set_formula,
        "set_range_values": _validate_set_range_values,
        "fill_formula": _validate_fill_formula,
        "set_value_if": _validate_set_value_if,
        "set_formula_if": _validate_set_formula_if,
        "draw_grid_border": _validate_draw_grid_border,
        "set_bold": _validate_set_bold,
        "set_fill_color": _validate_set_fill_color,
        "set_dimensions": _validate_set_dimensions,
        "merge_cells": _validate_merge_cells,
        "unmerge_cells": _validate_unmerge_cells,
        "set_alignment": _validate_set_alignment,
        "restore_design_snapshot": _validate_restore_design_snapshot,
    }
    return validators.get(op_type)


def _validate_add_sheet(op: PatchOp) -> None:
    """Validate add_sheet operation."""
    _validate_no_design_fields(op, op_name="add_sheet")
    if op.cell is not None:
        raise ValueError("add_sheet does not accept cell.")
    if op.range is not None:
        raise ValueError("add_sheet does not accept range.")
    if op.base_cell is not None:
        raise ValueError("add_sheet does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("add_sheet does not accept expected.")
    if op.value is not None:
        raise ValueError("add_sheet does not accept value.")
    if op.values is not None:
        raise ValueError("add_sheet does not accept values.")
    if op.formula is not None:
        raise ValueError("add_sheet does not accept formula.")


def _validate_cell_required(op: PatchOp) -> None:
    """Validate that the operation has a cell value."""
    if op.cell is None:
        raise ValueError(f"{op.op} requires cell.")


def _validate_set_value(op: PatchOp) -> None:
    """Validate set_value operation."""
    _validate_no_design_fields(op, op_name="set_value")
    if op.range is not None:
        raise ValueError("set_value does not accept range.")
    if op.base_cell is not None:
        raise ValueError("set_value does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("set_value does not accept expected.")
    if op.values is not None:
        raise ValueError("set_value does not accept values.")
    if op.formula is not None:
        raise ValueError("set_value does not accept formula.")


def _validate_set_formula(op: PatchOp) -> None:
    """Validate set_formula operation."""
    _validate_no_design_fields(op, op_name="set_formula")
    if op.range is not None:
        raise ValueError("set_formula does not accept range.")
    if op.base_cell is not None:
        raise ValueError("set_formula does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("set_formula does not accept expected.")
    if op.values is not None:
        raise ValueError("set_formula does not accept values.")
    if op.value is not None:
        raise ValueError("set_formula does not accept value.")
    if op.formula is None:
        raise ValueError("set_formula requires formula.")
    if not op.formula.startswith("="):
        raise ValueError("set_formula requires formula starting with '='.")


def _validate_set_range_values(op: PatchOp) -> None:
    """Validate set_range_values operation."""
    _validate_no_design_fields(op, op_name="set_range_values")
    if op.cell is not None:
        raise ValueError("set_range_values does not accept cell.")
    if op.base_cell is not None:
        raise ValueError("set_range_values does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("set_range_values does not accept expected.")
    if op.formula is not None:
        raise ValueError("set_range_values does not accept formula.")
    if op.range is None:
        raise ValueError("set_range_values requires range.")
    if op.values is None:
        raise ValueError("set_range_values requires values.")
    if not op.values:
        raise ValueError("set_range_values requires non-empty values.")
    if not all(op.values):
        raise ValueError("set_range_values values rows must not be empty.")
    expected_width = len(op.values[0])
    if any(len(row) != expected_width for row in op.values):
        raise ValueError("set_range_values requires rectangular values.")


def _validate_fill_formula(op: PatchOp) -> None:
    """Validate fill_formula operation."""
    _validate_no_design_fields(op, op_name="fill_formula")
    if op.cell is not None:
        raise ValueError("fill_formula does not accept cell.")
    if op.expected is not None:
        raise ValueError("fill_formula does not accept expected.")
    if op.value is not None:
        raise ValueError("fill_formula does not accept value.")
    if op.values is not None:
        raise ValueError("fill_formula does not accept values.")
    if op.range is None:
        raise ValueError("fill_formula requires range.")
    if op.base_cell is None:
        raise ValueError("fill_formula requires base_cell.")
    if op.formula is None:
        raise ValueError("fill_formula requires formula.")
    if not op.formula.startswith("="):
        raise ValueError("fill_formula requires formula starting with '='.")


def _validate_set_value_if(op: PatchOp) -> None:
    """Validate set_value_if operation."""
    _validate_no_design_fields(op, op_name="set_value_if")
    if op.formula is not None:
        raise ValueError("set_value_if does not accept formula.")
    if op.range is not None:
        raise ValueError("set_value_if does not accept range.")
    if op.values is not None:
        raise ValueError("set_value_if does not accept values.")
    if op.base_cell is not None:
        raise ValueError("set_value_if does not accept base_cell.")


def _validate_set_formula_if(op: PatchOp) -> None:
    """Validate set_formula_if operation."""
    _validate_no_design_fields(op, op_name="set_formula_if")
    if op.value is not None:
        raise ValueError("set_formula_if does not accept value.")
    if op.range is not None:
        raise ValueError("set_formula_if does not accept range.")
    if op.values is not None:
        raise ValueError("set_formula_if does not accept values.")
    if op.base_cell is not None:
        raise ValueError("set_formula_if does not accept base_cell.")
    if op.formula is None:
        raise ValueError("set_formula_if requires formula.")
    if not op.formula.startswith("="):
        raise ValueError("set_formula_if requires formula starting with '='.")


def _validate_draw_grid_border(op: PatchOp) -> None:
    """Validate draw_grid_border operation."""
    _validate_no_legacy_edit_fields(op, op_name="draw_grid_border")
    if op.cell is not None or op.range is not None:
        raise ValueError("draw_grid_border does not accept cell or range.")
    if op.bold is not None or op.fill_color is not None:
        raise ValueError("draw_grid_border does not accept bold or fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("draw_grid_border does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("draw_grid_border does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("draw_grid_border does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="draw_grid_border")
    if op.base_cell is None:
        raise ValueError("draw_grid_border requires base_cell.")
    if op.row_count is None or op.col_count is None:
        raise ValueError("draw_grid_border requires row_count and col_count.")
    if op.row_count < 1 or op.col_count < 1:
        raise ValueError("draw_grid_border requires row_count >= 1 and col_count >= 1.")
    if op.row_count * op.col_count > _MAX_STYLE_TARGET_CELLS:
        raise ValueError(
            f"draw_grid_border target exceeds max cells: {_MAX_STYLE_TARGET_CELLS}."
        )


def _validate_set_bold(op: PatchOp) -> None:
    """Validate set_bold operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_bold")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_bold does not accept row_count or col_count.")
    if op.fill_color is not None:
        raise ValueError("set_bold does not accept fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_bold does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_bold does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_bold does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_bold")
    _validate_exactly_one_cell_or_range(op, op_name="set_bold")
    if op.bold is None:
        op.bold = True
    _validate_style_target_size(op, op_name="set_bold")


def _validate_set_fill_color(op: PatchOp) -> None:
    """Validate set_fill_color operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_fill_color")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_fill_color does not accept row_count or col_count.")
    if op.bold is not None:
        raise ValueError("set_fill_color does not accept bold.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_fill_color does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_fill_color does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_fill_color does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_fill_color")
    _validate_exactly_one_cell_or_range(op, op_name="set_fill_color")
    if op.fill_color is None:
        raise ValueError("set_fill_color requires fill_color.")
    _validate_style_target_size(op, op_name="set_fill_color")


def _validate_set_dimensions(op: PatchOp) -> None:
    """Validate set_dimensions operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_dimensions")
    if op.cell is not None or op.range is not None or op.base_cell is not None:
        raise ValueError("set_dimensions does not accept cell/range/base_cell.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_dimensions does not accept row_count or col_count.")
    if op.bold is not None or op.fill_color is not None:
        raise ValueError("set_dimensions does not accept bold or fill_color.")
    if op.design_snapshot is not None:
        raise ValueError("set_dimensions does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_dimensions")
    has_rows = op.rows is not None
    has_columns = op.columns is not None
    if not has_rows and not has_columns:
        raise ValueError("set_dimensions requires rows and/or columns.")
    if has_rows and op.row_height is None:
        raise ValueError("set_dimensions requires row_height when rows is provided.")
    if has_columns and op.column_width is None:
        raise ValueError(
            "set_dimensions requires column_width when columns is provided."
        )
    if op.row_height is not None and op.row_height <= 0:
        raise ValueError("set_dimensions row_height must be > 0.")
    if op.column_width is not None and op.column_width <= 0:
        raise ValueError("set_dimensions column_width must be > 0.")


def _validate_merge_cells(op: PatchOp) -> None:
    """Validate merge_cells operation."""
    _validate_no_legacy_edit_fields(op, op_name="merge_cells")
    if op.cell is not None or op.base_cell is not None:
        raise ValueError("merge_cells does not accept cell or base_cell.")
    if op.range is None:
        raise ValueError("merge_cells requires range.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("merge_cells does not accept row_count or col_count.")
    if op.bold is not None or op.fill_color is not None:
        raise ValueError("merge_cells does not accept bold or fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("merge_cells does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("merge_cells does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("merge_cells does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="merge_cells")
    if _range_cell_count(op.range) < 2:
        raise ValueError("merge_cells requires a multi-cell range.")


def _validate_unmerge_cells(op: PatchOp) -> None:
    """Validate unmerge_cells operation."""
    _validate_no_legacy_edit_fields(op, op_name="unmerge_cells")
    if op.cell is not None or op.base_cell is not None:
        raise ValueError("unmerge_cells does not accept cell or base_cell.")
    if op.range is None:
        raise ValueError("unmerge_cells requires range.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("unmerge_cells does not accept row_count or col_count.")
    if op.bold is not None or op.fill_color is not None:
        raise ValueError("unmerge_cells does not accept bold or fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("unmerge_cells does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("unmerge_cells does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("unmerge_cells does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="unmerge_cells")


def _validate_set_alignment(op: PatchOp) -> None:
    """Validate set_alignment operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_alignment")
    if op.base_cell is not None:
        raise ValueError("set_alignment does not accept base_cell.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_alignment does not accept row_count or col_count.")
    if op.bold is not None or op.fill_color is not None:
        raise ValueError("set_alignment does not accept bold or fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_alignment does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_alignment does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_alignment does not accept design_snapshot.")
    _validate_exactly_one_cell_or_range(op, op_name="set_alignment")
    if (
        op.horizontal_align is None
        and op.vertical_align is None
        and op.wrap_text is None
    ):
        raise ValueError(
            "set_alignment requires at least one of horizontal_align, vertical_align, or wrap_text."
        )
    _validate_style_target_size(op, op_name="set_alignment")


def _validate_restore_design_snapshot(op: PatchOp) -> None:
    """Validate restore_design_snapshot operation."""
    _validate_no_legacy_edit_fields(op, op_name="restore_design_snapshot")
    if op.cell is not None or op.range is not None or op.base_cell is not None:
        raise ValueError(
            "restore_design_snapshot does not accept cell/range/base_cell."
        )
    if op.row_count is not None or op.col_count is not None:
        raise ValueError(
            "restore_design_snapshot does not accept row_count or col_count."
        )
    if op.bold is not None or op.fill_color is not None:
        raise ValueError("restore_design_snapshot does not accept bold or fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("restore_design_snapshot does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError(
            "restore_design_snapshot does not accept row_height or column_width."
        )
    _validate_no_alignment_fields(op, op_name="restore_design_snapshot")
    if op.design_snapshot is None:
        raise ValueError("restore_design_snapshot requires design_snapshot.")


def _validate_no_legacy_edit_fields(op: PatchOp, *, op_name: str) -> None:
    """Reject fields that are unrelated to design operations."""
    if op.expected is not None:
        raise ValueError(f"{op_name} does not accept expected.")
    if op.value is not None:
        raise ValueError(f"{op_name} does not accept value.")
    if op.values is not None:
        raise ValueError(f"{op_name} does not accept values.")
    if op.formula is not None:
        raise ValueError(f"{op_name} does not accept formula.")


def _validate_no_design_fields(op: PatchOp, *, op_name: str) -> None:
    """Reject design-only fields for legacy value edit operations."""
    if op.row_count is not None or op.col_count is not None:
        raise ValueError(f"{op_name} does not accept row_count or col_count.")
    if op.bold is not None:
        raise ValueError(f"{op_name} does not accept bold.")
    if op.fill_color is not None:
        raise ValueError(f"{op_name} does not accept fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError(f"{op_name} does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError(f"{op_name} does not accept row_height or column_width.")
    _validate_no_alignment_fields(op, op_name=op_name)
    if op.design_snapshot is not None:
        raise ValueError(f"{op_name} does not accept design_snapshot.")


def _validate_no_alignment_fields(op: PatchOp, *, op_name: str) -> None:
    """Reject alignment-only fields for unrelated operations."""
    if op.horizontal_align is not None:
        raise ValueError(f"{op_name} does not accept horizontal_align.")
    if op.vertical_align is not None:
        raise ValueError(f"{op_name} does not accept vertical_align.")
    if op.wrap_text is not None:
        raise ValueError(f"{op_name} does not accept wrap_text.")


def _validate_exactly_one_cell_or_range(op: PatchOp, *, op_name: str) -> None:
    """Ensure exactly one of cell/range is provided."""
    if op.base_cell is not None:
        raise ValueError(f"{op_name} does not accept base_cell.")
    has_cell = op.cell is not None
    has_range = op.range is not None
    if has_cell == has_range:
        raise ValueError(f"{op_name} requires exactly one of cell or range.")


def _validate_style_target_size(op: PatchOp, *, op_name: str) -> None:
    """Guard style edits against accidental huge targets."""
    target_count = 1 if op.cell is not None else _range_cell_count(op.range)
    if target_count > _MAX_STYLE_TARGET_CELLS:
        raise ValueError(
            f"{op_name} target exceeds max cells: {_MAX_STYLE_TARGET_CELLS}."
        )


def _range_cell_count(range_ref: str | None) -> int:
    """Return the number of cells represented by an A1 range."""
    if range_ref is None:
        raise ValueError("range is required.")
    start, end = range_ref.split(":", maxsplit=1)
    start_col, start_row = _split_a1(start)
    end_col, end_row = _split_a1(end)
    min_col = min(_column_label_to_index(start_col), _column_label_to_index(end_col))
    max_col = max(_column_label_to_index(start_col), _column_label_to_index(end_col))
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)
    return (max_col - min_col + 1) * (max_row - min_row + 1)


def _split_a1(value: str) -> tuple[str, int]:
    """Split A1 notation into normalized (column_label, row_index)."""
    if not _A1_PATTERN.match(value):
        raise ValueError(f"Invalid cell reference: {value}")
    idx = 0
    for index, char in enumerate(value):
        if char.isdigit():
            idx = index
            break
    column = value[:idx].upper()
    row = int(value[idx:])
    return column, row


def _normalize_column_identifier(value: str | int) -> str | int:
    """Normalize a column identifier preserving letter/index semantics."""
    if isinstance(value, int):
        if value < 1:
            raise ValueError("columns numeric values must be positive.")
        return value
    label = value.strip().upper()
    if not _COLUMN_LABEL_PATTERN.match(label):
        raise ValueError(f"Invalid column identifier: {value}")
    return label


def _column_label_to_index(label: str) -> int:
    """Convert Excel-style column label (A/AA) to 1-based index."""
    if not _COLUMN_LABEL_PATTERN.match(label):
        raise ValueError(f"Invalid column label: {label}")
    index = 0
    for char in label:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def _column_index_to_label(index: int) -> str:
    """Convert 1-based column index to Excel-style column label."""
    if index < 1:
        raise ValueError("Column index must be positive.")
    chunks: list[str] = []
    current = index
    while current > 0:
        current -= 1
        chunks.append(chr(ord("A") + (current % 26)))
        current //= 26
    return "".join(reversed(chunks))


class PatchValue(BaseModel):
    """Normalized before/after value in patch diff."""

    kind: PatchValueKind
    value: str | int | float | None


class PatchDiffItem(BaseModel):
    """Applied change record for patch operations."""

    op_index: int
    op: PatchOpType
    sheet: str
    cell: str | None = None
    before: PatchValue | None = None
    after: PatchValue | None = None
    status: PatchStatus = "applied"


class PatchErrorDetail(BaseModel):
    """Structured error details for patch failures."""

    op_index: int
    op: PatchOpType
    sheet: str
    cell: str | None
    message: str


class FormulaIssue(BaseModel):
    """Formula health-check finding."""

    sheet: str
    cell: str
    level: FormulaIssueLevel
    code: FormulaIssueCode
    message: str


class PatchRequest(BaseModel):
    """Input model for ExStruct MCP patch."""

    xlsx_path: Path
    ops: list[PatchOp]
    out_dir: Path | None = None
    out_name: str | None = None
    on_conflict: OnConflictPolicy = "overwrite"
    auto_formula: bool = False
    dry_run: bool = False
    return_inverse_ops: bool = False
    preflight_formula_check: bool = False


class PatchResult(BaseModel):
    """Output model for ExStruct MCP patch."""

    out_path: str
    patch_diff: list[PatchDiffItem] = Field(default_factory=list)
    inverse_ops: list[PatchOp] = Field(default_factory=list)
    formula_issues: list[FormulaIssue] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    error: PatchErrorDetail | None = None


def run_patch(
    request: PatchRequest, *, policy: PathPolicy | None = None
) -> PatchResult:
    """Run a patch operation and write the updated workbook.

    Args:
        request: Patch request payload.
        policy: Optional path policy for access control.

    Returns:
        Patch result with output path and diff.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If validation fails or the path violates policy.
        RuntimeError: If a backend operation fails.
    """
    resolved_input = _resolve_input_path(request.xlsx_path, policy=policy)
    _ensure_supported_extension(resolved_input)
    output_path = _resolve_output_path(
        resolved_input,
        out_dir=request.out_dir,
        out_name=request.out_name,
        policy=policy,
    )
    output_path, warning, skipped = _apply_conflict_policy(
        output_path, request.on_conflict
    )
    warnings: list[str] = []
    if warning:
        warnings.append(warning)
    if skipped and not request.dry_run:
        return PatchResult(
            out_path=str(output_path),
            patch_diff=[],
            inverse_ops=[],
            formula_issues=[],
            warnings=warnings,
        )
    if skipped and request.dry_run:
        warnings.append(
            "Dry-run mode ignores on_conflict=skip and simulates patch without writing."
        )

    com = get_com_availability()
    if resolved_input.suffix.lower() == ".xls" and not com.available:
        raise ValueError(
            ".xls editing requires Windows Excel COM (xlwings) in this environment."
        )
    if resolved_input.suffix.lower() == ".xls" and _contains_design_ops(request.ops):
        raise ValueError(
            "Design operations are not supported for .xls files. Convert to .xlsx/.xlsm first."
        )

    use_openpyxl = _requires_openpyxl_backend(request)
    if use_openpyxl and com.available:
        warnings.append("Using openpyxl backend for extended patch features.")

    _ensure_output_dir(output_path)
    if com.available and not use_openpyxl:
        try:
            diff = _apply_ops_xlwings(
                resolved_input,
                output_path,
                request.ops,
                request.auto_formula,
            )
            return PatchResult(
                out_path=str(output_path),
                patch_diff=diff,
                inverse_ops=[],
                formula_issues=[],
                warnings=warnings,
            )
        except PatchOpError as exc:
            return PatchResult(
                out_path=str(output_path),
                patch_diff=[],
                inverse_ops=[],
                formula_issues=[],
                warnings=warnings,
                error=exc.detail,
            )
        except Exception as exc:
            fallback = _maybe_fallback_openpyxl(
                request,
                resolved_input,
                output_path,
                warnings,
                reason=f"COM patch failed; falling back to openpyxl. ({exc!r})",
            )
            if fallback is not None:
                return fallback
            raise RuntimeError(f"COM patch failed: {exc}") from exc

    if com.reason:
        warnings.append(f"COM unavailable: {com.reason}")
    return _apply_with_openpyxl(
        request,
        resolved_input,
        output_path,
        warnings,
    )


def _apply_with_openpyxl(
    request: PatchRequest,
    input_path: Path,
    output_path: Path,
    warnings: list[str],
) -> PatchResult:
    """Apply patch operations using openpyxl."""
    try:
        diff, inverse_ops, formula_issues, op_warnings = _apply_ops_openpyxl(
            request,
            input_path,
            output_path,
        )
    except PatchOpError as exc:
        return PatchResult(
            out_path=str(output_path),
            patch_diff=[],
            inverse_ops=[],
            formula_issues=[],
            warnings=warnings,
            error=exc.detail,
        )
    except ValueError:
        raise
    except FileNotFoundError:
        raise
    except OSError:
        raise
    except Exception as exc:
        raise RuntimeError(f"openpyxl patch failed: {exc}") from exc

    warnings.extend(op_warnings)
    if not request.dry_run:
        warnings.append(
            "openpyxl editing may drop shapes/charts or unsupported elements."
        )
    _append_skip_warnings(warnings, diff)
    if (
        not request.dry_run
        and request.preflight_formula_check
        and any(issue.level == "error" for issue in formula_issues)
    ):
        issue = formula_issues[0]
        op_index, op_name = _find_preflight_issue_origin(issue, request.ops)
        error = PatchErrorDetail(
            op_index=op_index,
            op=op_name,
            sheet=issue.sheet,
            cell=issue.cell,
            message=f"Formula health check failed: {issue.message}",
        )
        return PatchResult(
            out_path=str(output_path),
            patch_diff=[],
            inverse_ops=[],
            formula_issues=formula_issues,
            warnings=warnings,
            error=error,
        )
    return PatchResult(
        out_path=str(output_path),
        patch_diff=diff,
        inverse_ops=inverse_ops,
        formula_issues=formula_issues,
        warnings=warnings,
    )


def _append_skip_warnings(warnings: list[str], diff: list[PatchDiffItem]) -> None:
    """Append warning messages for skipped conditional operations."""
    for item in diff:
        if item.status != "skipped":
            continue
        warnings.append(
            f"Skipped op[{item.op_index}] {item.op} at {item.sheet}!{item.cell} due to condition mismatch."
        )


def _find_preflight_issue_origin(
    issue: FormulaIssue, ops: list[PatchOp]
) -> tuple[int, PatchOpType]:
    """Find the most likely op index/op name for a preflight formula issue."""
    for index, op in enumerate(ops):
        if _op_targets_issue_cell(op, issue.sheet, issue.cell):
            return index, op.op
    return -1, "set_value"


def _op_targets_issue_cell(op: PatchOp, sheet: str, cell: str) -> bool:
    """Return True when an op can affect the specified sheet/cell."""
    if op.sheet != sheet:
        return False
    if op.cell is not None:
        return op.cell == cell
    if op.range is None:
        return False
    for row in _expand_range_coordinates(op.range):
        if cell in row:
            return True
    return False


def _maybe_fallback_openpyxl(
    request: PatchRequest,
    input_path: Path,
    output_path: Path,
    warnings: list[str],
    *,
    reason: str,
) -> PatchResult | None:
    """Attempt openpyxl fallback after COM failure."""
    if input_path.suffix.lower() == ".xls":
        warnings.append(reason)
        return None
    warnings.append(reason)
    return _apply_with_openpyxl(
        request,
        input_path,
        output_path,
        warnings,
    )


def _requires_openpyxl_backend(request: PatchRequest) -> bool:
    """Return True if request requires openpyxl backend for extended features."""
    if request.dry_run or request.return_inverse_ops or request.preflight_formula_check:
        return True
    return any(
        op.op
        in {
            "set_range_values",
            "fill_formula",
            "set_value_if",
            "set_formula_if",
            "draw_grid_border",
            "set_bold",
            "set_fill_color",
            "set_dimensions",
            "merge_cells",
            "unmerge_cells",
            "set_alignment",
            "restore_design_snapshot",
        }
        for op in request.ops
    )


def _contains_design_ops(ops: list[PatchOp]) -> bool:
    """Return True when any style/dimension design operation is present."""
    design_ops = {
        "draw_grid_border",
        "set_bold",
        "set_fill_color",
        "set_dimensions",
        "merge_cells",
        "unmerge_cells",
        "set_alignment",
        "restore_design_snapshot",
    }
    return any(op.op in design_ops for op in ops)


def _resolve_input_path(path: Path, *, policy: PathPolicy | None) -> Path:
    """Resolve and validate the input path."""
    resolved = policy.ensure_allowed(path) if policy else path.resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"Input file not found: {resolved}")
    if not resolved.is_file():
        raise ValueError(f"Input path is not a file: {resolved}")
    return resolved


def _ensure_supported_extension(path: Path) -> None:
    """Validate that the input file extension is supported."""
    if path.suffix.lower() not in _ALLOWED_EXTENSIONS:
        raise ValueError(f"Unsupported file extension: {path.suffix}")


def _resolve_output_path(
    input_path: Path,
    *,
    out_dir: Path | None,
    out_name: str | None,
    policy: PathPolicy | None,
) -> Path:
    """Build and validate the output path."""
    target_dir = out_dir or input_path.parent
    target_dir = policy.ensure_allowed(target_dir) if policy else target_dir.resolve()
    name = _normalize_output_name(input_path, out_name)
    output_path = (target_dir / name).resolve()
    if policy is not None:
        output_path = policy.ensure_allowed(output_path)
    return output_path


def _normalize_output_name(input_path: Path, out_name: str | None) -> str:
    """Normalize output filename with a safe suffix."""
    if out_name:
        candidate = Path(out_name)
        return (
            candidate.name
            if candidate.suffix
            else f"{candidate.name}{input_path.suffix}"
        )
    return f"{input_path.stem}_patched{input_path.suffix}"


def _ensure_output_dir(path: Path) -> None:
    """Ensure the output directory exists before writing."""
    path.parent.mkdir(parents=True, exist_ok=True)


def _apply_conflict_policy(
    output_path: Path, on_conflict: OnConflictPolicy
) -> tuple[Path, str | None, bool]:
    """Apply output conflict policy to a resolved output path."""
    if not output_path.exists():
        return output_path, None, False
    if on_conflict == "skip":
        return (
            output_path,
            f"Output exists; skipping write: {output_path.name}",
            True,
        )
    if on_conflict == "rename":
        renamed = _next_available_path(output_path)
        return (
            renamed,
            f"Output exists; renamed to: {renamed.name}",
            False,
        )
    return output_path, None, False


def _next_available_path(path: Path) -> Path:
    """Return the next available path by appending a numeric suffix."""
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for idx in range(1, 10_000):
        candidate = path.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Failed to resolve unique path for {path}")


def _apply_ops_openpyxl(
    request: PatchRequest,
    input_path: Path,
    output_path: Path,
) -> tuple[list[PatchDiffItem], list[PatchOp], list[FormulaIssue], list[str]]:
    """Apply operations using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    if input_path.suffix.lower() == ".xls":
        raise ValueError("openpyxl cannot edit .xls files.")

    if input_path.suffix.lower() == ".xlsm":
        workbook = load_workbook(input_path, keep_vba=True)
    else:
        workbook = load_workbook(input_path)
    try:
        diff, inverse_ops, op_warnings = _apply_ops_to_openpyxl_workbook(
            workbook,
            request.ops,
            request.auto_formula,
            return_inverse_ops=request.return_inverse_ops,
        )
        formula_issues = (
            _collect_formula_issues_openpyxl(workbook)
            if request.preflight_formula_check
            else []
        )
        if not request.dry_run and not (
            request.preflight_formula_check
            and any(issue.level == "error" for issue in formula_issues)
        ):
            workbook.save(output_path)
    finally:
        workbook.close()
    return diff, inverse_ops, formula_issues, op_warnings


def _apply_ops_to_openpyxl_workbook(
    workbook: OpenpyxlWorkbookProtocol,
    ops: list[PatchOp],
    auto_formula: bool,
    *,
    return_inverse_ops: bool,
) -> tuple[list[PatchDiffItem], list[PatchOp], list[str]]:
    """Apply ops to an openpyxl workbook instance."""
    sheets = _openpyxl_sheet_map(workbook)
    diff: list[PatchDiffItem] = []
    inverse_ops: list[PatchOp] = []
    op_warnings: list[str] = []
    for index, op in enumerate(ops):
        try:
            item, inverse = _apply_openpyxl_op(
                workbook, sheets, op, index, auto_formula, op_warnings
            )
            diff.append(item)
            if return_inverse_ops and item.status == "applied" and inverse is not None:
                inverse_ops.append(inverse)
        except ValueError as exc:
            raise PatchOpError.from_op(index, op, exc) from exc
    if return_inverse_ops:
        inverse_ops.reverse()
    return diff, inverse_ops, op_warnings


def _openpyxl_sheet_map(
    workbook: OpenpyxlWorkbookProtocol,
) -> dict[str, OpenpyxlWorksheetProtocol]:
    """Build a sheet map for openpyxl workbooks."""
    sheet_names = getattr(workbook, "sheetnames", None)
    if not isinstance(sheet_names, list):
        raise ValueError("Invalid workbook: sheetnames missing.")
    return {name: workbook[name] for name in sheet_names}


def _apply_openpyxl_op(
    workbook: OpenpyxlWorkbookProtocol,
    sheets: dict[str, OpenpyxlWorksheetProtocol],
    op: PatchOp,
    index: int,
    auto_formula: bool,
    warnings: list[str],
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply a single op to openpyxl workbook."""
    if op.op == "add_sheet":
        return _apply_openpyxl_add_sheet(workbook, sheets, op, index)

    existing_sheet = sheets.get(op.sheet)
    if existing_sheet is None:
        raise ValueError(f"Sheet not found: {op.sheet}")
    return _apply_openpyxl_sheet_op(
        existing_sheet,
        op,
        index,
        auto_formula=auto_formula,
        warnings=warnings,
    )


def _apply_openpyxl_sheet_op(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
    *,
    auto_formula: bool,
    warnings: list[str],
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply openpyxl operation that targets an existing sheet."""
    if op.op in {"set_value", "set_formula", "set_value_if", "set_formula_if"}:
        return _apply_openpyxl_cell_op(sheet, op, index, auto_formula)
    handlers: dict[PatchOpType, Callable[[], tuple[PatchDiffItem, PatchOp | None]]] = {
        "set_range_values": lambda: _apply_openpyxl_set_range_values(sheet, op, index),
        "fill_formula": lambda: _apply_openpyxl_fill_formula(sheet, op, index),
        "draw_grid_border": lambda: _apply_openpyxl_draw_grid_border(sheet, op, index),
        "set_bold": lambda: _apply_openpyxl_set_bold(sheet, op, index),
        "set_fill_color": lambda: _apply_openpyxl_set_fill_color(sheet, op, index),
        "set_dimensions": lambda: _apply_openpyxl_set_dimensions(sheet, op, index),
        "merge_cells": lambda: _apply_openpyxl_merge_cells(sheet, op, index, warnings),
        "unmerge_cells": lambda: _apply_openpyxl_unmerge_cells(sheet, op, index),
        "set_alignment": lambda: _apply_openpyxl_set_alignment(sheet, op, index),
        "restore_design_snapshot": lambda: _apply_openpyxl_restore_design_snapshot(
            sheet, op, index
        ),
    }
    handler = handlers.get(op.op)
    if handler is None:
        raise ValueError(f"Unsupported op: {op.op}")
    return handler()


def _apply_openpyxl_add_sheet(
    workbook: OpenpyxlWorkbookProtocol,
    sheets: dict[str, OpenpyxlWorksheetProtocol],
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply add_sheet op."""
    if op.sheet in sheets:
        raise ValueError(f"Sheet already exists: {op.sheet}")
    sheet = workbook.create_sheet(title=op.sheet)
    sheets[op.sheet] = sheet
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="sheet", value=op.sheet),
        ),
        None,
    )


def _apply_openpyxl_set_range_values(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_range_values op."""
    if op.range is None or op.values is None:
        raise ValueError("set_range_values requires range and values.")
    coordinates = _expand_range_coordinates(op.range)
    rows, cols = _shape_of_coordinates(coordinates)
    if len(op.values) != rows:
        raise ValueError("set_range_values values height does not match range.")
    if any(len(row) != cols for row in op.values):
        raise ValueError("set_range_values values width does not match range.")
    for r_idx, row in enumerate(coordinates):
        for c_idx, coord in enumerate(row):
            sheet[coord].value = op.values[r_idx][c_idx]
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="value", value=f"{rows}x{cols}"),
        ),
        None,
    )


def _apply_openpyxl_fill_formula(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply fill_formula op."""
    if op.range is None or op.formula is None or op.base_cell is None:
        raise ValueError("fill_formula requires range, base_cell and formula.")
    coordinates = _expand_range_coordinates(op.range)
    rows, cols = _shape_of_coordinates(coordinates)
    if rows != 1 and cols != 1:
        raise ValueError("fill_formula range must be a single row or a single column.")
    for row in coordinates:
        for coord in row:
            sheet[coord].value = _translate_formula(op.formula, op.base_cell, coord)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="formula", value=op.formula),
        ),
        None,
    )


def _apply_openpyxl_draw_grid_border(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply draw_grid_border op with thin black border."""
    if op.base_cell is None or op.row_count is None or op.col_count is None:
        raise ValueError(
            "draw_grid_border requires base_cell, row_count and col_count."
        )
    coordinates = _expand_rect_coordinates(op.base_cell, op.row_count, op.col_count)
    snapshot = DesignSnapshot(
        borders=[_snapshot_border(sheet[coord], coord) for coord in coordinates]
    )
    for coord in coordinates:
        _set_grid_border(sheet[coord])
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=f"{op.base_cell}:{coordinates[-1]}",
            before=None,
            after=PatchValue(kind="style", value="grid_border(thin,black)"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_bold(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_bold op."""
    targets = _resolve_style_targets(op)
    target_bold = True if op.bold is None else op.bold
    snapshot = DesignSnapshot(
        fonts=[_snapshot_font(sheet[coord], coord) for coord in targets]
    )
    for coord in targets:
        cell = sheet[coord]
        font = copy(cell.font)
        font.bold = target_bold
        cell.font = font
    location = op.cell if op.cell is not None else op.range
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=f"bold={target_bold}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_fill_color(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_fill_color op."""
    if op.fill_color is None:
        raise ValueError("set_fill_color requires fill_color.")
    try:
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    targets = _resolve_style_targets(op)
    snapshot = DesignSnapshot(
        fills=[_snapshot_fill(sheet[coord], coord) for coord in targets]
    )
    normalized = _normalize_hex_color(op.fill_color)
    for coord in targets:
        sheet[coord].fill = PatternFill(
            fill_type="solid",
            start_color=normalized,
            end_color=normalized,
        )
    location = op.cell if op.cell is not None else op.range
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=f"fill={normalized}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_dimensions(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_dimensions op."""
    snapshot = DesignSnapshot()
    parts: list[str] = []
    if op.rows is not None and op.row_height is not None:
        for row in op.rows:
            row_dimension = sheet.row_dimensions[row]
            snapshot.row_dimensions.append(
                RowDimensionSnapshot(
                    row=row,
                    height=getattr(row_dimension, "height", None),
                )
            )
            row_dimension.height = op.row_height
        parts.append(f"rows={len(op.rows)}")
    if op.columns is not None and op.column_width is not None:
        normalized_columns = _normalize_columns_for_dimensions(op.columns)
        for column in normalized_columns:
            column_dimension = sheet.column_dimensions[column]
            snapshot.column_dimensions.append(
                ColumnDimensionSnapshot(
                    column=column,
                    width=getattr(column_dimension, "width", None),
                )
            )
            column_dimension.width = op.column_width
        parts.append(f"columns={len(normalized_columns)}")
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="dimension", value=", ".join(parts)),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_merge_cells(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
    warnings: list[str],
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply merge_cells op."""
    if op.range is None:
        raise ValueError("merge_cells requires range.")
    overlapped = _intersecting_merged_ranges(sheet, op.range)
    if overlapped:
        raise ValueError(
            "merge_cells range overlaps existing merged ranges: "
            + ", ".join(overlapped)
            + "."
        )
    merge_warning = _build_merge_value_loss_warning(sheet, op.sheet, op.range)
    if merge_warning is not None:
        warnings.append(merge_warning)
    snapshot = DesignSnapshot(
        merge_state=MergeStateSnapshot(scope=op.range, ranges=[]),
    )
    sheet.merge_cells(op.range)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="style", value=f"merged={op.range}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_unmerge_cells(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply unmerge_cells op."""
    if op.range is None:
        raise ValueError("unmerge_cells requires range.")
    target_ranges = _intersecting_merged_ranges(sheet, op.range)
    snapshot = DesignSnapshot(
        merge_state=MergeStateSnapshot(scope=op.range, ranges=target_ranges),
    )
    for range_ref in target_ranges:
        sheet.unmerge_cells(range_ref)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="style", value=f"unmerged={len(target_ranges)}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_alignment(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_alignment op."""
    targets = _resolve_style_targets(op)
    snapshot = DesignSnapshot(
        alignments=[_snapshot_alignment(sheet[coord], coord) for coord in targets]
    )
    for coord in targets:
        cell = sheet[coord]
        alignment = copy(cell.alignment)
        if op.horizontal_align is not None:
            alignment.horizontal = op.horizontal_align
        if op.vertical_align is not None:
            alignment.vertical = op.vertical_align
        if op.wrap_text is not None:
            alignment.wrap_text = op.wrap_text
        cell.alignment = alignment
    location = op.cell if op.cell is not None else op.range
    summary = (
        f"horizontal={op.horizontal_align},"
        f"vertical={op.vertical_align},"
        f"wrap_text={op.wrap_text}"
    )
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=summary),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_restore_design_snapshot(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply restore_design_snapshot op."""
    if op.design_snapshot is None:
        raise ValueError("restore_design_snapshot requires design_snapshot.")
    _restore_design_snapshot(sheet, op.design_snapshot)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="style", value="design_snapshot_restored"),
        ),
        None,
    )


def _apply_openpyxl_cell_op(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
    auto_formula: bool,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply single-cell operations."""
    cell_ref = op.cell
    if cell_ref is None:
        raise ValueError(f"{op.op} requires cell.")
    cell = sheet[cell_ref]
    before = _openpyxl_cell_value(cell)

    if op.op == "set_value":
        after = _set_cell_value(cell, op.value, auto_formula, op_name="set_value")
        return _build_cell_result(
            op, index, cell_ref, before, after
        ), _build_inverse_cell_op(op, cell_ref, before)
    if op.op == "set_formula":
        formula = _require_formula(op.formula, "set_formula")
        cell.value = formula
        after = PatchValue(kind="formula", value=formula)
        return _build_cell_result(
            op, index, cell_ref, before, after
        ), _build_inverse_cell_op(op, cell_ref, before)
    if op.op == "set_value_if":
        if not _values_equal_for_condition(
            _patch_value_to_primitive(before), op.expected
        ):
            return _build_skipped_result(op, index, cell_ref, before), None
        after = _set_cell_value(cell, op.value, auto_formula, op_name="set_value_if")
        return _build_cell_result(
            op, index, cell_ref, before, after
        ), _build_inverse_cell_op(op, cell_ref, before)
    formula_if = _require_formula(op.formula, "set_formula_if")
    if not _values_equal_for_condition(_patch_value_to_primitive(before), op.expected):
        return _build_skipped_result(op, index, cell_ref, before), None
    cell.value = formula_if
    after = PatchValue(kind="formula", value=formula_if)
    return _build_cell_result(
        op, index, cell_ref, before, after
    ), _build_inverse_cell_op(op, cell_ref, before)


def _set_cell_value(
    cell: OpenpyxlCellProtocol,
    value: str | int | float | None,
    auto_formula: bool,
    *,
    op_name: str,
) -> PatchValue:
    """Set cell value with auto_formula handling."""
    if isinstance(value, str) and value.startswith("="):
        if not auto_formula:
            raise ValueError(f"{op_name} rejects values starting with '='.")
        cell.value = value
        return PatchValue(kind="formula", value=value)
    cell.value = value
    return PatchValue(kind="value", value=value)


def _build_cell_result(
    op: PatchOp,
    index: int,
    cell_ref: str,
    before: PatchValue | None,
    after: PatchValue | None,
) -> PatchDiffItem:
    """Build applied diff item for single-cell op."""
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=cell_ref,
        before=before,
        after=after,
    )


def _build_skipped_result(
    op: PatchOp,
    index: int,
    cell_ref: str,
    before: PatchValue | None,
) -> PatchDiffItem:
    """Build skipped diff item."""
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=cell_ref,
        before=before,
        after=before,
        status="skipped",
    )


def _require_formula(formula: str | None, op_name: str) -> str:
    """Require a non-null formula string."""
    if formula is None:
        raise ValueError(f"{op_name} requires formula.")
    return formula


def _openpyxl_cell_value(cell: OpenpyxlCellProtocol) -> PatchValue | None:
    """Normalize an openpyxl cell value into PatchValue."""
    value = getattr(cell, "value", None)
    if value is None:
        return None
    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        text = _normalize_formula(value)
        return PatchValue(kind="formula", value=text)
    return PatchValue(kind="value", value=value)


def _normalize_formula(value: object) -> str:
    """Ensure formula string starts with '='."""
    text = str(value)
    return text if text.startswith("=") else f"={text}"


def _expand_range_coordinates(range_ref: str) -> list[list[str]]:
    """Expand A1 range string into a 2D list of coordinates."""
    try:
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    if min_col > max_col or min_row > max_row:
        raise ValueError(f"Invalid range reference: {range_ref}")
    rows: list[list[str]] = []
    for row_idx in range(min_row, max_row + 1):
        row: list[str] = []
        for col_idx in range(min_col, max_col + 1):
            row.append(f"{get_column_letter(col_idx)}{row_idx}")
        rows.append(row)
    return rows


def _shape_of_coordinates(coordinates: list[list[str]]) -> tuple[int, int]:
    """Return rows/cols for expanded coordinates."""
    if not coordinates or not coordinates[0]:
        raise ValueError("Range expansion resulted in an empty coordinate set.")
    return len(coordinates), len(coordinates[0])


def _expand_rect_coordinates(base_cell: str, rows: int, cols: int) -> list[str]:
    """Expand base cell + size into a flat coordinate list."""
    base_column, base_row = _split_a1(base_cell)
    start_col = _column_label_to_index(base_column)
    coordinates: list[str] = []
    for row_offset in range(rows):
        for col_offset in range(cols):
            column = _column_index_to_label(start_col + col_offset)
            coordinates.append(f"{column}{base_row + row_offset}")
    return coordinates


def _resolve_style_targets(op: PatchOp) -> list[str]:
    """Resolve style operation target coordinates."""
    if op.cell is not None:
        return [op.cell]
    if op.range is None:
        raise ValueError(f"{op.op} requires cell or range.")
    coordinates = _expand_range_coordinates(op.range)
    targets: list[str] = []
    for row in coordinates:
        targets.extend(row)
    return targets


def _merged_range_strings(sheet: OpenpyxlWorksheetProtocol) -> list[str]:
    """Return normalized merged range strings from worksheet."""
    merged_cells = getattr(sheet, "merged_cells", None)
    ranges = getattr(merged_cells, "ranges", None)
    if ranges is None:
        return []
    return [str(item) for item in ranges]


def _intersecting_merged_ranges(
    sheet: OpenpyxlWorksheetProtocol, scope_range: str
) -> list[str]:
    """Return merged ranges that intersect the scope."""
    intersections: list[str] = []
    for merged_range in _merged_range_strings(sheet):
        if _ranges_overlap(scope_range, merged_range):
            intersections.append(merged_range)
    return intersections


def _ranges_overlap(left: str, right: str) -> bool:
    """Return True if two A1 ranges overlap."""
    left_min_col, left_min_row, left_max_col, left_max_row = _range_bounds(left)
    right_min_col, right_min_row, right_max_col, right_max_row = _range_bounds(right)
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def _range_bounds(range_ref: str) -> tuple[int, int, int, int]:
    """Return range boundaries in (min_col, min_row, max_col, max_row)."""
    try:
        from openpyxl.utils.cell import range_boundaries
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    return cast(tuple[int, int, int, int], range_boundaries(range_ref))


def _build_merge_value_loss_warning(
    sheet: OpenpyxlWorksheetProtocol,
    sheet_name: str,
    range_ref: str,
) -> str | None:
    """Build warning when merge can clear non-top-left cell values."""
    coordinates = _expand_range_coordinates(range_ref)
    top_left = coordinates[0][0]
    risky_cells: list[str] = []
    for row in coordinates:
        for coord in row:
            if coord == top_left:
                continue
            value = sheet[coord].value
            if _has_non_empty_cell_value(value):
                risky_cells.append(coord)
    if not risky_cells:
        return None
    joined = ", ".join(risky_cells)
    return (
        f"merge_cells may clear non-top-left values at {sheet_name}!{range_ref}: "
        f"{joined}"
    )


def _has_non_empty_cell_value(value: str | int | float | None) -> bool:
    """Return True when cell has a non-empty value."""
    if value is None:
        return False
    if isinstance(value, str):
        return value != ""
    return True


def _normalize_hex_color(fill_color: str) -> str:
    """Normalize #RRGGBB/#AARRGGBB into AARRGGBB."""
    text = fill_color.strip().upper()
    if not _HEX_COLOR_PATTERN.match(text):
        raise ValueError("Invalid fill_color format. Use '#RRGGBB' or '#AARRGGBB'.")
    raw = text[1:]
    return raw if len(raw) == 8 else f"FF{raw}"


def _normalize_columns_for_dimensions(columns: list[str | int]) -> list[str]:
    """Normalize columns list to unique Excel-style labels."""
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in columns:
        label = (
            _column_index_to_label(raw) if isinstance(raw, int) else raw.strip().upper()
        )
        if label in seen:
            continue
        seen.add(label)
        normalized.append(label)
    return normalized


def _set_grid_border(cell: OpenpyxlCellProtocol) -> None:
    """Set thin black border on all sides."""
    try:
        from openpyxl.styles import Side
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    side = Side(style="thin", color="FF000000")
    border = copy(cell.border)
    border.top = side
    border.right = side
    border.bottom = side
    border.left = side
    cell.border = border


def _snapshot_border(cell: OpenpyxlCellProtocol, coordinate: str) -> BorderSnapshot:
    """Capture border snapshot for one cell."""
    border = cell.border
    return BorderSnapshot(
        cell=coordinate,
        top=_snapshot_border_side(border.top),
        right=_snapshot_border_side(border.right),
        bottom=_snapshot_border_side(border.bottom),
        left=_snapshot_border_side(border.left),
    )


def _snapshot_border_side(side: object) -> BorderSideSnapshot:
    """Capture one border side state."""
    style = getattr(side, "style", None)
    color = _extract_openpyxl_color(getattr(side, "color", None))
    return BorderSideSnapshot(style=style, color=color)


def _snapshot_font(cell: OpenpyxlCellProtocol, coordinate: str) -> FontSnapshot:
    """Capture font snapshot for one cell."""
    font = cell.font
    return FontSnapshot(cell=coordinate, bold=getattr(font, "bold", None))


def _snapshot_fill(cell: OpenpyxlCellProtocol, coordinate: str) -> FillSnapshot:
    """Capture fill snapshot for one cell."""
    fill = cell.fill
    return FillSnapshot(
        cell=coordinate,
        fill_type=getattr(fill, "fill_type", None),
        start_color=_extract_openpyxl_color(getattr(fill, "start_color", None)),
        end_color=_extract_openpyxl_color(getattr(fill, "end_color", None)),
    )


def _snapshot_alignment(
    cell: OpenpyxlCellProtocol, coordinate: str
) -> AlignmentSnapshot:
    """Capture alignment snapshot for one cell."""
    alignment = cell.alignment
    return AlignmentSnapshot(
        cell=coordinate,
        horizontal=getattr(alignment, "horizontal", None),
        vertical=getattr(alignment, "vertical", None),
        wrap_text=getattr(alignment, "wrap_text", None),
    )


def _extract_openpyxl_color(color: object) -> str | None:
    """Extract RGB-like color text from openpyxl color object."""
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return None
    text = str(rgb).upper()
    return text if len(text) == 8 else None


def _build_restore_snapshot_op(sheet: str, snapshot: DesignSnapshot) -> PatchOp | None:
    """Build a restore op when snapshot contains data."""
    if (
        not snapshot.borders
        and not snapshot.fonts
        and not snapshot.fills
        and not snapshot.alignments
        and snapshot.merge_state is None
        and not snapshot.row_dimensions
        and not snapshot.column_dimensions
    ):
        return None
    return PatchOp(op="restore_design_snapshot", sheet=sheet, design_snapshot=snapshot)


def _restore_design_snapshot(
    sheet: OpenpyxlWorksheetProtocol,
    snapshot: DesignSnapshot,
) -> None:
    """Restore cell style and dimension snapshot."""
    if snapshot.merge_state is not None:
        _restore_merge_state(sheet, snapshot.merge_state)
    for border_snapshot in snapshot.borders:
        _restore_border(sheet[border_snapshot.cell], border_snapshot)
    for font_snapshot in snapshot.fonts:
        cell = sheet[font_snapshot.cell]
        font = copy(cell.font)
        font.bold = font_snapshot.bold
        cell.font = font
    for fill_snapshot in snapshot.fills:
        _restore_fill(sheet[fill_snapshot.cell], fill_snapshot)
    for alignment_snapshot in snapshot.alignments:
        _restore_alignment(sheet[alignment_snapshot.cell], alignment_snapshot)
    for row_snapshot in snapshot.row_dimensions:
        sheet.row_dimensions[row_snapshot.row].height = row_snapshot.height
    for column_snapshot in snapshot.column_dimensions:
        sheet.column_dimensions[column_snapshot.column].width = column_snapshot.width


def _restore_merge_state(
    sheet: OpenpyxlWorksheetProtocol,
    snapshot: MergeStateSnapshot,
) -> None:
    """Restore merged ranges for a scope deterministically."""
    for range_ref in _intersecting_merged_ranges(sheet, snapshot.scope):
        sheet.unmerge_cells(range_ref)
    for range_ref in snapshot.ranges:
        sheet.merge_cells(range_ref)


def _restore_border(cell: OpenpyxlCellProtocol, snapshot: BorderSnapshot) -> None:
    """Restore border from snapshot."""
    border = copy(cell.border)
    border.top = _build_side_from_snapshot(snapshot.top)
    border.right = _build_side_from_snapshot(snapshot.right)
    border.bottom = _build_side_from_snapshot(snapshot.bottom)
    border.left = _build_side_from_snapshot(snapshot.left)
    cell.border = border


def _build_side_from_snapshot(snapshot: BorderSideSnapshot) -> OpenpyxlSideProtocol:
    """Build openpyxl Side object from serializable snapshot."""
    try:
        from openpyxl.styles import Side
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    kwargs: dict[str, str] = {}
    if snapshot.style is not None:
        kwargs["style"] = snapshot.style
    if snapshot.color is not None:
        kwargs["color"] = snapshot.color
    return cast(OpenpyxlSideProtocol, Side(**kwargs))


def _restore_fill(cell: OpenpyxlCellProtocol, snapshot: FillSnapshot) -> None:
    """Restore fill from snapshot."""
    try:
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    cell.fill = PatternFill(
        fill_type=snapshot.fill_type,
        start_color=snapshot.start_color,
        end_color=snapshot.end_color,
    )


def _restore_alignment(cell: OpenpyxlCellProtocol, snapshot: AlignmentSnapshot) -> None:
    """Restore alignment from snapshot."""
    alignment = copy(cell.alignment)
    alignment.horizontal = snapshot.horizontal
    alignment.vertical = snapshot.vertical
    alignment.wrap_text = snapshot.wrap_text
    cell.alignment = alignment


def _translate_formula(formula: str, origin: str, target: str) -> str:
    """Translate formula with relative references from origin to target."""
    try:
        from openpyxl.formula.translate import Translator
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    translated = Translator(formula, origin=origin).translate_formula(target)
    return str(translated)


def _patch_value_to_primitive(value: PatchValue | None) -> str | int | float | None:
    """Convert PatchValue into primitive value for condition checks."""
    if value is None:
        return None
    return value.value


def _values_equal_for_condition(
    current: str | int | float | None,
    expected: str | int | float | None,
) -> bool:
    """Compare values for conditional update checks."""
    return current == expected


def _build_inverse_cell_op(
    op: PatchOp,
    cell_ref: str,
    before: PatchValue | None,
) -> PatchOp | None:
    """Build inverse operation for single-cell updates."""
    if op.op not in {"set_value", "set_formula", "set_value_if", "set_formula_if"}:
        return None
    if before is None:
        return PatchOp(op="set_value", sheet=op.sheet, cell=cell_ref, value=None)
    if before.kind == "formula":
        return PatchOp(
            op="set_formula",
            sheet=op.sheet,
            cell=cell_ref,
            formula=str(before.value),
        )
    return PatchOp(op="set_value", sheet=op.sheet, cell=cell_ref, value=before.value)


def _collect_formula_issues_openpyxl(
    workbook: OpenpyxlWorkbookProtocol,
) -> list[FormulaIssue]:
    """Collect simple formula issues by scanning formula text."""
    token_map: dict[str, tuple[FormulaIssueCode, FormulaIssueLevel]] = {
        "#REF!": ("ref_error", "error"),
        "#NAME?": ("name_error", "error"),
        "#DIV/0!": ("div0_error", "error"),
        "#VALUE!": ("value_error", "error"),
        "#N/A": ("na_error", "warning"),
    }
    issues: list[FormulaIssue] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        iter_rows = getattr(sheet, "iter_rows", None)
        if iter_rows is None:
            continue
        for row in iter_rows():
            for cell in row:
                raw = getattr(cell, "value", None)
                if not isinstance(raw, str) or not raw.startswith("="):
                    continue
                normalized = raw.upper()
                if "==" in normalized:
                    issues.append(
                        FormulaIssue(
                            sheet=sheet_name,
                            cell=str(getattr(cell, "coordinate", "")),
                            level="warning",
                            code="invalid_token",
                            message="Formula contains duplicated '=' token.",
                        )
                    )
                for token, (code, level) in token_map.items():
                    if token in normalized:
                        issues.append(
                            FormulaIssue(
                                sheet=sheet_name,
                                cell=str(getattr(cell, "coordinate", "")),
                                level=level,
                                code=code,
                                message=f"Formula contains error token {token}.",
                            )
                        )
    return issues


def _apply_ops_xlwings(
    input_path: Path,
    output_path: Path,
    ops: list[PatchOp],
    auto_formula: bool,
) -> list[PatchDiffItem]:
    """Apply operations using Excel COM via xlwings."""
    diff: list[PatchDiffItem] = []
    try:
        with _xlwings_workbook(input_path) as workbook:
            sheets = {sheet.name: sheet for sheet in workbook.sheets}
            for index, op in enumerate(ops):
                try:
                    diff.append(
                        _apply_xlwings_op(workbook, sheets, op, index, auto_formula)
                    )
                except ValueError as exc:
                    raise PatchOpError.from_op(index, op, exc) from exc
            workbook.save(str(output_path))
    except ValueError:
        raise
    except Exception as exc:
        raise RuntimeError(f"COM patch failed: {exc}") from exc
    return diff


def _apply_xlwings_op(
    workbook: XlwingsWorkbookProtocol,
    sheets: dict[str, XlwingsSheetProtocol],
    op: PatchOp,
    index: int,
    auto_formula: bool,
) -> PatchDiffItem:
    """Apply a single op to an xlwings workbook."""
    # Extended ops are routed to openpyxl by _requires_openpyxl_backend.
    # Keep this explicit guard to prevent accidental path regressions.
    if op.op not in {"add_sheet", "set_value", "set_formula"}:
        raise ValueError(f"Unsupported op: {op.op}")
    if op.op == "add_sheet":
        if op.sheet in sheets:
            raise ValueError(f"Sheet already exists: {op.sheet}")
        last = workbook.sheets[-1] if workbook.sheets else None
        sheet = workbook.sheets.add(name=op.sheet, after=last)
        sheets[op.sheet] = sheet
        return PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="sheet", value=op.sheet),
        )

    existing_sheet = sheets.get(op.sheet)
    if existing_sheet is None:
        raise ValueError(f"Sheet not found: {op.sheet}")
    cell_ref = op.cell
    if cell_ref is None:
        raise ValueError(f"{op.op} requires cell.")
    rng = existing_sheet.range(cell_ref)
    before = _xlwings_cell_value(rng)
    if op.op == "set_value":
        if isinstance(op.value, str) and op.value.startswith("="):
            if not auto_formula:
                raise ValueError("set_value rejects values starting with '='.")
            rng.formula = op.value
            after = PatchValue(kind="formula", value=op.value)
        else:
            rng.value = op.value
            after = PatchValue(kind="value", value=op.value)
        return PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=cell_ref,
            before=before,
            after=after,
        )
    if op.op == "set_formula":
        formula = op.formula
        if formula is None:
            raise ValueError("set_formula requires formula.")
        rng.formula = formula
        after = PatchValue(kind="formula", value=formula)
        return PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=cell_ref,
            before=before,
            after=after,
        )
    raise ValueError(f"Unsupported op: {op.op}")


def _xlwings_cell_value(cell: XlwingsRangeProtocol) -> PatchValue | None:
    """Normalize an xlwings cell value into PatchValue."""
    formula = getattr(cell, "formula", None)
    if isinstance(formula, str) and formula.startswith("="):
        return PatchValue(kind="formula", value=formula)
    value = getattr(cell, "value", None)
    if value is None:
        return None
    return PatchValue(kind="value", value=value)


@contextmanager
def _xlwings_workbook(file_path: Path) -> Iterator[XlwingsWorkbookProtocol]:
    """Open an Excel workbook with a dedicated COM app."""
    app = xw.App(add_book=False, visible=False)
    app.display_alerts = False
    app.screen_updating = False
    workbook = app.books.open(str(file_path))
    try:
        yield workbook
    finally:
        try:
            workbook.close()
        except Exception:
            pass
        try:
            app.quit()
        except Exception:
            try:
                app.kill()
            except Exception:
                pass


class PatchOpError(ValueError):
    """Patch operation error with structured detail."""

    def __init__(self, detail: PatchErrorDetail) -> None:
        super().__init__(detail.message)
        self.detail = detail

    @classmethod
    def from_op(cls, index: int, op: PatchOp, exc: Exception) -> PatchOpError:
        """Build a PatchOpError from an op and exception."""
        detail = PatchErrorDetail(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.cell,
            message=str(exc),
        )
        return cls(detail)
