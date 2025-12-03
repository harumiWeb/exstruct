import logging

logger = logging.getLogger(__name__)
_warned_keys = set()

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import numpy as np
from pathlib import Path
from typing import List, Tuple


def warn_once(key: str, message: str):
    if key not in _warned_keys:
        logger.warning(message)
        _warned_keys.add(key)


def load_border_maps_xlsx(xlsx_path: Path, sheet_name: str):
    # スタイル取得のため read_only=False にする（重要）
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Sheet '{sheet_name}' not found in {xlsx_path}")

    ws = wb[sheet_name]

    # used range の外接矩形に限定（速度向上）
    # calculate_dimension が 'A1:F40' のような文字列を返す
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ws.calculate_dimension())
    except Exception:
        # フォールバック（空シートなど）
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1

    # +1 は 1-based を直接インデックスに使うための余白（[row][col] でアクセス）
    shape = (max_row + 1, max_col + 1)
    has_border = np.zeros(shape, dtype=bool)
    top_edge = np.zeros(shape, dtype=bool)
    bottom_edge = np.zeros(shape, dtype=bool)
    left_edge = np.zeros(shape, dtype=bool)
    right_edge = np.zeros(shape, dtype=bool)

    def edge_has_style(edge) -> bool:
        # edge.style が None なら「罫線なし」
        # 場合によって 'none' 文字列が来ることもあるため両方除外
        if edge is None:
            return False
        style = getattr(edge, "style", None)
        return style is not None and style != "none"

    # r/c はインデックスから決定し、ws.cell でセルを取得
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            b = getattr(cell, "border", None)
            if b is None:
                continue

            t = edge_has_style(b.top)
            btm = edge_has_style(b.bottom)
            l = edge_has_style(b.left)
            rgt = edge_has_style(b.right)

            if t or btm or l or rgt:
                has_border[r, c] = True
                if t:
                    top_edge[r, c] = True
                if btm:
                    bottom_edge[r, c] = True
                if l:
                    left_edge[r, c] = True
                if rgt:
                    right_edge[r, c] = True

    wb.close()
    return has_border, top_edge, bottom_edge, left_edge, right_edge, max_row, max_col


def detect_border_clusters(has_border: np.ndarray, min_size: int = 4):
    try:
        from scipy.ndimage import label

        structure = np.array([[0, 1, 0], [1, 1, 1], [0, 1, 0]], dtype=np.uint8)
        lbl, num = label(has_border.astype(np.uint8), structure=structure)  # type: ignore
        rects: List[Tuple[int, int, int, int]] = []
        for k in range(1, num + 1):
            ys, xs = np.where(lbl == k)
            if len(ys) < min_size:
                continue
            rects.append((int(ys.min()), int(xs.min()), int(ys.max()), int(xs.max())))
        return rects
    except Exception:
        warn_once(
            "scipy-missing",
            "scipy is not available. Falling back to pure-Python BFS for connected components, which may be significantly slower.",
        )
        # scipyが無い場合はBFS
        from collections import deque

        h, w = has_border.shape
        visited = np.zeros_like(has_border, dtype=bool)
        rects: List[Tuple[int, int, int, int]] = []
        for r in range(h):
            for c in range(w):
                if not has_border[r, c] or visited[r, c]:
                    continue
                q = deque([(r, c)])
                visited[r, c] = True
                ys = [r]
                xs = [c]
                while q:
                    yy, xx = q.popleft()
                    for dy, dx in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                        ny, nx = yy + dy, xx + dx
                        if (
                            0 <= ny < h
                            and 0 <= nx < w
                            and has_border[ny, nx]
                            and not visited[ny, nx]
                        ):
                            visited[ny, nx] = True
                            q.append((ny, nx))
                            ys.append(ny)
                            xs.append(nx)
                if len(ys) >= min_size:
                    rects.append((min(ys), min(xs), max(ys), max(xs)))
        return rects


def _get_values_block(ws, top, left, bottom, right):
    # 2D配列（Pythonのlist of list）として値を取得
    vals = []
    for row in ws.iter_rows(
        min_row=top, max_row=bottom, min_col=left, max_col=right, values_only=True
    ):
        vals.append(list(row))
    return vals


def shrink_to_content_openpyxl(
    ws,
    top: int,
    left: int,
    bottom: int,
    right: int,
    require_inside_border: bool,
    top_edge,
    bottom_edge,
    left_edge,
    right_edge,
    min_nonempty_ratio: float = 0.0,
) -> Tuple[int, int, int, int]:
    vals = _get_values_block(ws, top, left, bottom, right)
    rows_n = bottom - top + 1
    cols_n = right - left + 1

    def to_str(x):
        return "" if x is None else str(x)

    def is_empty_value(x):
        return to_str(x).strip() == ""

    def row_nonempty_ratio_local(i: int) -> float:
        if cols_n <= 0:
            return 0.0
        row = vals[i]
        cnt = sum(1 for v in row if not is_empty_value(v))
        return cnt / cols_n

    def col_nonempty_ratio_local(j: int) -> float:
        if rows_n <= 0:
            return 0.0
        cnt = 0
        for i in range(rows_n):
            if not is_empty_value(vals[i][j]):
                cnt += 1
        return cnt / rows_n

    def col_has_inside_border(j_abs: int) -> bool:
        if not require_inside_border:
            return False
        count_pairs = 0
        for r_abs in range(top, bottom + 1):
            if (
                j_abs > left
                and right_edge[r_abs, j_abs - 1]
                and left_edge[r_abs, j_abs]
            ):
                count_pairs += 1
        return count_pairs > 0

    def row_has_inside_border(i_abs: int) -> bool:
        if not require_inside_border:
            return False
        count_pairs = 0
        for c_abs in range(left, right + 1):
            if i_abs > top and bottom_edge[i_abs - 1, c_abs] and top_edge[i_abs, c_abs]:
                count_pairs += 1
        return count_pairs > 0

    # 左
    while left <= right and cols_n > 0:
        empty_col = all(
            not (
                top_edge[i, left]
                or bottom_edge[i, left]
                or left_edge[i, left]
                or right_edge[i, left]
            )
            for i in range(top, bottom + 1)
        )
        if (
            empty_col
            or (require_inside_border and not col_has_inside_border(left))
            or (
                min_nonempty_ratio > 0.0
                and col_nonempty_ratio_local(0) < min_nonempty_ratio
            )
        ):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(0)
            cols_n -= 1
            left += 1
        else:
            break
    # 上
    while top <= bottom and rows_n > 0:
        empty_row = all(
            not (
                top_edge[top, j]
                or bottom_edge[top, j]
                or left_edge[top, j]
                or right_edge[top, j]
            )
            for j in range(left, right + 1)
        )
        if (
            empty_row
            or (require_inside_border and not row_has_inside_border(top))
            or (
                min_nonempty_ratio > 0.0
                and row_nonempty_ratio_local(0) < min_nonempty_ratio
            )
        ):
            vals.pop(0)
            rows_n -= 1
            top += 1
        else:
            break
    # 右
    while left <= right and cols_n > 0:
        empty_col = all(
            not (
                top_edge[i, right]
                or bottom_edge[i, right]
                or left_edge[i, right]
                or right_edge[i, right]
            )
            for i in range(top, bottom + 1)
        )
        if (
            empty_col
            or (require_inside_border and not col_has_inside_border(right))
            or (
                min_nonempty_ratio > 0.0
                and col_nonempty_ratio_local(cols_n - 1) < min_nonempty_ratio
            )
        ):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(cols_n - 1)
            cols_n -= 1
            right -= 1
        else:
            break
    # 下
    while top <= bottom and rows_n > 0:
        empty_row = all(
            not (
                top_edge[bottom, j]
                or bottom_edge[bottom, j]
                or left_edge[bottom, j]
                or right_edge[bottom, j]
            )
            for j in range(left, right + 1)
        )
        if (
            empty_row
            or (require_inside_border and not row_has_inside_border(bottom))
            or (
                min_nonempty_ratio > 0.0
                and row_nonempty_ratio_local(rows_n - 1) < min_nonempty_ratio
            )
        ):
            vals.pop(rows_n - 1)
            rows_n -= 1
            bottom -= 1
        else:
            break
    return top, left, bottom, right
