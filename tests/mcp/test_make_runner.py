from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from exstruct.cli.availability import ComAvailability
from exstruct.mcp import patch_runner
from exstruct.mcp.io import PathPolicy
from exstruct.mcp.patch_runner import MakeRequest, PatchOp, run_make


def _disable_com(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )


def _write_workbook(path: Path, value: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = value
    workbook.save(path)
    workbook.close()


def test_run_make_creates_xlsx_with_sheet1(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    result = run_make(
        MakeRequest(out_path=out_path, ops=[]),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    assert result.engine == "openpyxl"
    workbook = load_workbook(result.out_path)
    try:
        assert "Sheet1" in workbook.sheetnames
    finally:
        workbook.close()


def test_run_make_applies_ops(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[
                PatchOp(op="add_sheet", sheet="Data"),
                PatchOp(op="set_value", sheet="Data", cell="A1", value="ok"),
            ],
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        assert workbook["Data"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_run_make_conflict_overwrite(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    _write_workbook(out_path, "old")
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="overwrite",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(out_path)
    try:
        assert workbook["Sheet1"]["A1"].value == "new"
    finally:
        workbook.close()


def test_run_make_conflict_skip(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    _write_workbook(out_path, "old")
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="skip",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.patch_diff == []
    workbook = load_workbook(out_path)
    try:
        assert workbook["Sheet1"]["A1"].value == "old"
    finally:
        workbook.close()


def test_run_make_conflict_rename(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    _write_workbook(out_path, "old")
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert Path(result.out_path) != out_path
    assert Path(result.out_path).exists()


def test_run_make_rejects_path_outside_root(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    root = tmp_path / "root"
    root.mkdir()
    with pytest.raises(ValueError):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xlsx"),
            policy=PathPolicy(root=root),
        )


def test_run_make_rejects_xls_when_com_unavailable(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    with pytest.raises(ValueError, match=r"requires Windows Excel COM"):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xls"),
            policy=PathPolicy(root=tmp_path),
        )


def test_run_make_rejects_xls_with_openpyxl_backend(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match=r"backend='openpyxl' cannot edit \.xls files"):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xls", backend="openpyxl"),
            policy=PathPolicy(root=tmp_path),
        )


def test_run_make_rejects_xls_with_dry_run_flags(tmp_path: Path) -> None:
    with pytest.raises(
        ValueError,
        match=r"\.xls creation does not support dry_run, return_inverse_ops, or preflight_formula_check",
    ):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xls", dry_run=True),
            policy=PathPolicy(root=tmp_path),
        )
