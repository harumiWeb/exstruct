from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
import pytest

from exstruct.cli.availability import ComAvailability
from exstruct.mcp import patch_runner
from exstruct.mcp.io import PathPolicy
from exstruct.mcp.patch_runner import PatchOp, PatchRequest, run_patch


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "old"
    sheet["B1"] = 1
    workbook.save(path)
    workbook.close()


def _disable_com(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )


def test_run_patch_set_value_and_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new"),
        PatchOp(op="set_formula", sheet="Sheet1", cell="B1", formula="=SUM(1,1)"),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert Path(result.out_path).exists()
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A1"].value == "new"
        formula_value = sheet["B1"].value
        if isinstance(formula_value, str) and not formula_value.startswith("="):
            formula_value = f"={formula_value}"
        assert formula_value == "=SUM(1,1)"
    finally:
        workbook.close()
    assert len(result.patch_diff) == 2
    assert result.patch_diff[0].after is not None


def test_run_patch_add_sheet_and_set_value(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="add_sheet", sheet="NewSheet"),
        PatchOp(op="set_value", sheet="NewSheet", cell="A1", value="ok"),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    workbook = load_workbook(result.out_path)
    try:
        assert "NewSheet" in workbook.sheetnames
        assert workbook["NewSheet"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_run_patch_add_sheet_rejects_duplicate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="add_sheet", sheet="Sheet1")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert result.error.op_index == 0


def test_patch_op_allows_value_starting_with_equal() -> None:
    op = PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="=SUM(1,1)")
    assert op.value == "=SUM(1,1)"


def test_run_patch_rejects_formula_without_equal() -> None:
    with pytest.raises(ValidationError):
        PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="SUM(1,1)")


def test_run_patch_set_value_with_equal_requires_auto_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="=SUM(1,1)")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert "rejects values starting with" in result.error.message


def test_run_patch_set_value_with_equal_auto_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="=SUM(1,1)")]
    request = PatchRequest(
        xlsx_path=input_path, ops=ops, on_conflict="rename", auto_formula=True
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    workbook = load_workbook(result.out_path)
    try:
        formula_value = workbook["Sheet1"]["A1"].value
        if isinstance(formula_value, str) and not formula_value.startswith("="):
            formula_value = f"={formula_value}"
        assert formula_value == "=SUM(1,1)"
    finally:
        workbook.close()


def test_run_patch_rejects_path_outside_root(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    root = tmp_path / "root"
    root.mkdir()
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    with pytest.raises(ValueError):
        run_patch(request, policy=PathPolicy(root=root))


def test_run_patch_conflict_rename(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.out_path != str(default_out)
    assert Path(result.out_path).exists()


def test_run_patch_conflict_skip(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="skip")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.out_path == str(default_out)
    assert result.patch_diff == []
    assert any("skipping" in warning for warning in result.warnings)


def test_run_patch_conflict_skip_dry_run_still_simulates(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="skip",
        dry_run=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.patch_diff) == 1
    assert any("ignores on_conflict=skip" in warning for warning in result.warnings)
    assert not any("may drop shapes/charts" in warning for warning in result.warnings)
    assert default_out.read_text(encoding="utf-8") == "dummy"


def test_run_patch_conflict_overwrite(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    _create_workbook(default_out)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="overwrite")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.out_path == str(default_out)
    workbook = load_workbook(result.out_path)
    try:
        assert workbook["Sheet1"]["A1"].value == "new"
    finally:
        workbook.close()


def test_run_patch_atomicity(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x"),
        PatchOp(op="set_value", sheet="Missing", cell="A1", value="y"),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    output_path = tmp_path / "book_patched.xlsx"
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert not output_path.exists()


def test_run_patch_creates_out_dir(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    out_dir = tmp_path / "nested" / "output"
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        out_dir=out_dir,
        on_conflict="rename",
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    out_path = Path(result.out_path)
    assert out_path.exists()
    assert out_path.parent == out_dir


def test_run_patch_xls_requires_com(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xls"
    input_path.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="add_sheet", sheet="Sheet2")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    with pytest.raises(ValueError, match=r"requires Windows Excel COM"):
        run_patch(request, policy=PathPolicy(root=tmp_path))


def test_run_patch_xlsm_openpyxl_uses_keep_vba(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsm"
    input_path.write_bytes(b"dummy")
    calls: dict[str, object] = {}

    class FakeCell:
        def __init__(self, value: str | int | float | None) -> None:
            self.value = value
            self.data_type: str | None = None

    class FakeSheet:
        def __init__(self) -> None:
            self._cells: dict[str, FakeCell] = {"A1": FakeCell("old")}

        def __getitem__(self, key: str) -> FakeCell:
            if key not in self._cells:
                self._cells[key] = FakeCell(None)
            return self._cells[key]

    class FakeWorkbook:
        def __init__(self) -> None:
            self._sheets: dict[str, FakeSheet] = {"Sheet1": FakeSheet()}
            self.sheetnames = ["Sheet1"]

        def __getitem__(self, key: str) -> FakeSheet:
            return self._sheets[key]

        def create_sheet(self, title: str) -> FakeSheet:
            sheet = FakeSheet()
            self._sheets[title] = sheet
            self.sheetnames.append(title)
            return sheet

        def save(self, filename: str | Path) -> None:
            calls["saved"] = str(filename)

        def close(self) -> None:
            calls["closed"] = True

    fake_workbook = FakeWorkbook()

    def _fake_load_workbook(path: Path, **kwargs: object) -> FakeWorkbook:
        calls["path"] = str(path)
        calls["keep_vba"] = kwargs.get("keep_vba", False)
        return fake_workbook

    monkeypatch.setattr("openpyxl.load_workbook", _fake_load_workbook)

    request = PatchRequest(
        xlsx_path=input_path,
        ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
        on_conflict="rename",
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert calls["keep_vba"] is True
    assert calls["closed"] is True


def test_run_patch_dry_run_does_not_write(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    output_path = tmp_path / "book_patched.xlsx"
    assert not output_path.exists()
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        dry_run=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert not output_path.exists()
    assert len(result.patch_diff) == 1


def test_run_patch_return_inverse_ops(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        return_inverse_ops=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.inverse_ops) == 1
    inverse = result.inverse_ops[0]
    assert inverse.op == "set_value"
    assert inverse.cell == "A1"
    assert inverse.value == "old"


def test_run_patch_set_range_values(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="set_range_values",
            sheet="Sheet1",
            range="A2:B3",
            values=[["r1c1", "r1c2"], ["r2c1", "r2c2"]],
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A2"].value == "r1c1"
        assert sheet["B3"].value == "r2c2"
    finally:
        workbook.close()


def test_run_patch_set_range_values_size_mismatch(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="set_range_values",
            sheet="Sheet1",
            range="A2:B3",
            values=[["only_one_column"], ["still_one_column"]],
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert "width does not match range" in result.error.message


def test_run_patch_set_value_if_skipped(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="set_value_if",
            sheet="Sheet1",
            cell="A1",
            expected="not_old",
            value="new",
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.patch_diff) == 1
    assert result.patch_diff[0].status == "skipped"


def test_run_patch_fill_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        sheet = workbook["Sheet1"]
        sheet["A2"] = 1
        sheet["A3"] = 2
        sheet["A4"] = 3
        sheet["B2"] = 10
        sheet["B3"] = 20
        sheet["B4"] = 30
        workbook.save(input_path)
    finally:
        workbook.close()
    ops = [
        PatchOp(
            op="fill_formula",
            sheet="Sheet1",
            range="C2:C4",
            base_cell="C2",
            formula="=A2+B2",
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["C2"].value == "=A2+B2"
        assert sheet["C3"].value == "=A3+B3"
        assert sheet["C4"].value == "=A4+B4"
    finally:
        workbook.close()


def test_run_patch_formula_health_check(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="=#REF!+1")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        preflight_formula_check=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert result.formula_issues
    assert result.formula_issues[0].code == "ref_error"


def test_run_patch_formula_health_check_reports_matching_op(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_formula", sheet="Sheet1", cell="B1", formula="=SUM(1,1)"),
        PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="=#REF!+1"),
    ]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        preflight_formula_check=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert result.error.op_index == 1
    assert result.error.op == "set_formula"


def test_patch_op_add_sheet_rejects_unrelated_fields() -> None:
    with pytest.raises(ValidationError, match="add_sheet does not accept range"):
        PatchOp(op="add_sheet", sheet="NewSheet", range="A1:A1")


def test_patch_op_set_value_rejects_expected() -> None:
    with pytest.raises(ValidationError, match="set_value does not accept expected"):
        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x", expected="old")
